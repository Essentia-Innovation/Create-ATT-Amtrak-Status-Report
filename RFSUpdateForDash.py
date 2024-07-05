from openpyxl import load_workbook  #Library for working with excel files
from glob import glob #Library for finding files matching a pattern
import pandas #Library for data manipulation and analysis
import json #Library for working with JSON data
from requests.auth import HTTPBasicAuth #Library for handling HTTP Basic Auth
import requests #Library for making HTTP requests

# Global variable to verify if columns are correct
verifyColumns = False


class RFSUpdate:
    def __init__(self):
        # Load the settings from a JSON file containing API credentials
        with open('settings.json', "rb") as PFile:
            password_data = json.loads(PFile.read().decode('utf-8'))

            """Initialize ToeSpeed with necessary parameters for API"""
            # Assign variables

            self.url_onevizion = password_data["urlOneVizion"]
            self.login_onevizion = password_data["loginOneVizion"]
            self.pass_onevizion = password_data["passOneVizion"]
            # Create an HTTPBasicAuth object using the login and password for authentication
            self.auth_onevizion = HTTPBasicAuth(self.login_onevizion, self.pass_onevizion)
            # Define headers for API calls
            self.headers = {'Content-type': 'application/json', 'Content-Encoding': 'utf-8'}


        # Generate list of client dictionaries from eSpeed to compare against spreadsheet
        rfsdict = self.search_trackors("ENTProject", ["XITOR_KEY", "ENTPR_RFS_PHASE", "ENTPR_RFS_ACTIVE_TASK",
                                                             "ENTPR_NOTES_HISTORY"],
                                              'is_not_null(XITOR_KEY)')
        # Find the first Excel file in the directory
        file = glob("*.xlsx")[0]
        # Load the workbook and select the first worksheet
        wb = load_workbook(file)
        ws = wb.worksheets[0]
        # Convert the worksheet data to a pandas DataFrame
        sheet_df = pandas.DataFrame(ws.values)
        new_header = sheet_df.iloc[0]  # grab the first row for the header
        sheet_df = sheet_df[1:]  # Take the data excluding the header row
        sheet_df.columns = new_header  # set the header row as the dataframe header

        # Verify that the columns match the expected column names, 0 is column A, 8 is column I, 10 is column K, 11 is column L in the original Excel file.  We are using pandas here so the index counting starts at 0.
        if new_header[0] == 'RFS Number' and new_header[8] == 'Essentia-RFS Phase Update' and new_header[
            10] == 'Essentia-RFS Active  Task' and new_header[11] == 'Essentia ETA-Update':
            verifyColumns = True

        if verifyColumns:
            # Loop through each row in the worksheet.  Note below the row numbers are 1-based, not 0-based. So indexing starts at 1 because we are using the openpyxl library not pandas here.
            for rowNum in range(2, ws.max_row + 1):
                # Check if the cell in column 1 (RFS Number) is not None
                if ws.cell(rowNum, 1).value is not None:
                    # Loop through each dictionary in rfsdict
                    for i in rfsdict:
                        if ws.cell(rowNum, 1).value == i["TRACKOR_KEY"]: # Check if the RFS Number in the Excel file matches the TRACKOR_KEY in the dictionary
                            print("MATCH: ")
                            print("Espeed Data: ", i["TRACKOR_KEY"], "Phase: ", i["ENTPR_RFS_PHASE"], "Active Task: ",
                                  i["ENTPR_RFS_ACTIVE_TASK"], "Notes:", i["ENTPR_NOTES_HISTORY"]) # Print the data from the dictionary
                            print("Excel Data: ", ws.cell(rowNum, 1).value, "Phase: ", ws.cell(rowNum, 9).value,
                                  "Active Task: ", ws.cell(rowNum, 11).value, "Notes: ", ws.cell(rowNum, 12).value) # Print the data from the Excel file
                            ws.cell(rowNum, 12).value = i["ENTPR_NOTES_HISTORY"] # Update the Notes column in the Excel file with the Notes from the dictionary.  Column 12 is column L. 
                            if ws.cell(rowNum, 8).value != i["ENTPR_RFS_PHASE"] or ws.cell(rowNum, 10).value != i[
                                "ENTPR_RFS_ACTIVE_TASK"]: # Check if the Phase or Active Task in the Excel file do not match the Phase or Active Task in the dictionary. Column 8 is column H, Column 10 is column J.
                                ws.cell(rowNum, 9).value = i["ENTPR_RFS_PHASE"] # Update the Phase column in the Excel file with the Phase from the dictionary. Column 9 is column I.
                                ws.cell(rowNum, 11).value = i["ENTPR_RFS_ACTIVE_TASK"] # Update the Active Task column in the Excel file with the Active Task from the dictionary. Column 11 is column K.
                            print("New Excel Data: ", ws.cell(rowNum, 1).value, "Phase: ", ws.cell(rowNum, 9).value,
                                  "Active Task: ", ws.cell(rowNum, 11).value, "Notes: ", ws.cell(rowNum, 12).value) # Print the updated data from the Excel file
                            print("")

            wb.save('output.xlsx') # Save the updated Excel file
        else:
            print("Column names do not match expected columns.") # Print an error message if the columns do not match the expected column names

    def search_trackors(self, trackor_type: str, fields: list, filter: str):
            """Search for trackors through a filter using OneVizion API.

            Parameters
            ----------
            trackor_type : str
                Trackor Type of the Trackor for which data should be retrieved (e.g., BILL_OF_MATERIALS).
                Use the Trackor Type name, not the Trackor Type label.

            fields: list
                Comma separated list of Field names to return in response.
                Note: If Field doesn't exists HTTP 404 (Not found) will be returned

            filter: str
                Supported filter operators:
                equal(cf_name, value) or =(cf_name, value)
                greater(cf_name, value) or >(cf_name, value)
                less(cf_name, value) or <(cf_name, value)
                gt_today(cf_name, value = +0) or >=Today(cf_name, value = +0) [use + or - in front of value without spaces, if not specified +0 will be used]
                lt_today(cf_name, value = +0) or <=Today(cf_name, value = +0) [use + or - in front of value without spaces, if not specified +0 will be used]
                within(cf_name, value1, value2)
                not_equal(cf_name, value) or <>(cf_name, value)
                null(cf_name) or is_null(cf_name)
                is_not_null(cf_name)
                outer_equal(cf_name, value)
                outer_not_equal(cf_name, value)
                less_or_equal(cf_name, value) or <=(cf_name, value)
                greater_or_equal(cf_name, value) or >=(cf_name, value)
                this_week(cf_name, value = +0) [use + or - in front of value without spaces, if not specified +0 will be used]
                this_month(cf_name, value = +0) [use + or - in front of value without spaces, if not specified +0 will be used]
                this_quarter(cf_name, value = +0) [use + or - in front of value without spaces, if not specified +0 will be used] (equal to This FQ)
                this_year(cf_name, value = +0) [use + or - in front of value without spaces, if not specified +0 will be used] (equal to This FY)
                this_week_to_date(cf_name)
                this_month_to_date(cf_name)
                this_quarter_to_date(cf_name) (equal to This FQ to Dt)
                this_year_to_date(cf_name) (equal to This FY to Dt)
                field_equal(cf1_name, cf2_name) or =F(cf1_name, cf2_name)
                field_not_equal(cf1_name, cf2_name) or <>F(cf1_name, cf2_name)
                field_less(cf1_name, cf2_name) or <F(cf1_name, cf2_name)
                field_greater(cf1_name, cf2_name) or >F(cf1_name, cf2_name)
                field_less_or_equal(cf1_name, cf2_name) or <=F(cf1_name, cf2_name)
                field_greater_or_equal(cf1_name, cf2_name) or >=F(cf1_name, cf2_name)
                new(cf_name) or is_new(cf_name)
                not_new(cf_name) or is_not_new(cf_name)
                equal_myself(cf_name) or =Myself(cf_name)
                not_equal_myself(cf_name) or <>Myself(cf_name)

            Raises
            ------
            Exception
                If the response from onevizion is not ok
            """

            url = 'https://{url_onevizion}/api/v3/trackor_types/{trackor_type}/trackors/search?fields={fields}'.format(
                url_onevizion=self.url_onevizion,
                trackor_type=trackor_type, fields=",".join(fields)) # Create the URL for the API call
            answer = requests.post(url, data=filter, headers=self.headers, auth=self.auth_onevizion) # Make the API call
            if answer.ok:
                return answer.json() # Return the JSON response
            else:
                raise Exception(answer.text) # Raise an exception if the response is not ok

if __name__ == "__main__":
        RFSUpdate()